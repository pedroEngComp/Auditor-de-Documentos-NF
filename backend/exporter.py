import csv
import io
import json
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ─── Campos que vão para a planilha de resultados ─────────────────────────────

RESULT_FIELDS = [
    "filename",
    "numero_documento",
    "tipo_documento",
    "data_emissao",
    "data_emissao_nf",
    "data_pagamento",
    "fornecedor",
    "cnpj_fornecedor",
    "descricao_servico",
    "valor_bruto",
    "valor_bruto_num",
    "aprovado_por",
    "banco_destino",
    "status",
    "hash_verificacao",
    "parse_status",
    "encoding_used",
    "extraction_source",
    "anomaly_count",
    "anomaly_rules",
    "processed_at",
    "prompt_version",
]


def _flatten_anomalies(doc: dict) -> dict:

    anomalies = doc.get("anomalies", [])
    doc["anomaly_count"] = len(anomalies)
    doc["anomaly_rules"] = " | ".join(a["rule"] for a in anomalies)
    return doc


# ─── CSV ──────────────────────────────────────────────────────────────────────

def export_to_csv(docs: list[dict]) -> bytes:

    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=RESULT_FIELDS,
        extrasaction="ignore",
        lineterminator="\n",
    )
    writer.writeheader()
    for doc in docs:
        _flatten_anomalies(doc)
        writer.writerow({f: doc.get(f, "") for f in RESULT_FIELDS})

    # UTF-8 BOM faz o Excel reconhecer acentos automaticamente
    return b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")


# ─── Excel ────────────────────────────────────────────────────────────────────

def export_to_excel(docs: list[dict]) -> bytes:

    wb = openpyxl.Workbook()

    # ── Aba 1: Resultados ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resultados"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    anomaly_fill = PatternFill("solid", fgColor="FFE0E0")

    # Cabeçalho
    for col, field in enumerate(RESULT_FIELDS, start=1):
        cell = ws1.cell(row=1, column=col, value=field.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, doc in enumerate(docs, start=2):
        _flatten_anomalies(doc)
        has_anomaly = doc.get("anomaly_count", 0) > 0
        for col, field in enumerate(RESULT_FIELDS, start=1):
            cell = ws1.cell(row=row_idx, column=col, value=str(doc.get(field, "") or ""))
            if has_anomaly:
                cell.fill = anomaly_fill

    # Ajusta largura das colunas automaticamente
    for col in ws1.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── Aba 2: Anomalias detalhadas ───────────────────────────────────────────
    ws2 = wb.create_sheet("Anomalias")
    anomaly_headers = ["filename", "numero_documento", "fornecedor", "rule",
                       "description", "field_evidence", "confidence",
                       "processed_at", "prompt_version"]

    for col, h in enumerate(anomaly_headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h.upper())
        cell.fill = header_fill
        cell.font = header_font

    row_idx = 2
    for doc in docs:
        for anomaly in doc.get("anomalies", []):
            ws2.cell(row=row_idx, column=1, value=doc.get("filename", ""))
            ws2.cell(row=row_idx, column=2, value=doc.get("numero_documento", ""))
            ws2.cell(row=row_idx, column=3, value=doc.get("fornecedor", ""))
            ws2.cell(row=row_idx, column=4, value=anomaly.get("rule", ""))
            ws2.cell(row=row_idx, column=5, value=anomaly.get("description", ""))
            ws2.cell(row=row_idx, column=6, value=anomaly.get("field_evidence", ""))
            ws2.cell(row=row_idx, column=7, value=anomaly.get("confidence", ""))
            ws2.cell(row=row_idx, column=8, value=doc.get("processed_at", ""))
            ws2.cell(row=row_idx, column=9, value=doc.get("prompt_version", ""))
            row_idx += 1

    for col in ws2.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ─── Log de auditoria ─────────────────────────────────────────────────────────

def export_audit_log(docs: list[dict]) -> bytes:

    output = io.StringIO()
    log_fields = [
        "timestamp_log",
        "filename",
        "numero_documento",
        "parse_status",
        "extraction_source",
        "event_type",       # PROCESSED | ANOMALY_DETECTED | PARSE_ERROR
        "event_detail",
        "rule",
        "field_evidence",
        "confidence",
        "prompt_version",
        "processed_at",
    ]

    writer = csv.DictWriter(output, fieldnames=log_fields, lineterminator="\n")
    writer.writeheader()

    now = datetime.utcnow().isoformat()

    for doc in docs:
        base = {
            "timestamp_log": now,
            "filename": doc.get("filename", ""),
            "numero_documento": doc.get("numero_documento", ""),
            "parse_status": doc.get("parse_status", ""),
            "extraction_source": doc.get("extraction_source", ""),
            "prompt_version": doc.get("prompt_version", ""),
            "processed_at": doc.get("processed_at", ""),
        }

        # Registro de processamento do arquivo
        writer.writerow({
            **base,
            "event_type": "PARSE_ERROR" if "error" in (doc.get("parse_status") or "") else "PROCESSED",
            "event_detail": doc.get("parse_error") or f"Processado via {doc.get('extraction_source', 'local')}",
            "rule": "",
            "field_evidence": "",
            "confidence": "",
        })

        # Um registro por anomalia detectada
        for anomaly in doc.get("anomalies", []):
            writer.writerow({
                **base,
                "event_type": "ANOMALY_DETECTED",
                "event_detail": anomaly.get("description", ""),
                "rule": anomaly.get("rule", ""),
                "field_evidence": anomaly.get("field_evidence", ""),
                "confidence": anomaly.get("confidence", ""),
            })

    return b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")
